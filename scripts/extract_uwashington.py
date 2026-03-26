#!/usr/bin/env python3

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl
from pypdf import PdfReader


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DIR = ROOT / "College-Data" / "UWashington"
OUTPUT_PATH = ROOT / "src" / "data" / "schools" / "uwashington.json"

XLSX_FILES = {
    "2018-2019": "2018-2019-Common-Data-Set_Seattle_updated.xlsx",
    "2019-2020": "CDS_2019-2020_seattle.xlsx",
    "2020-2021": "CDS_2020-2021_Seattle.xlsx",
    "2021-2022": "CDS_2021-2022_Seattle.xlsx",
    "2022-2023": "CDS_2022-2023_Seattle.xlsx",
    "2023-2024": "CDS_Seattle_2023_2024.xlsx",
}

PDF_FILES = {
    "2024-2025": "CDS_2024-2025_Seattle.pdf",
    "2025-2026": "CDS_2025-2026_Seattle.pdf",
}


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def compact_row(values: list[Any]) -> list[Any]:
    compacted: list[Any] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            stripped = value.strip()
            if stripped:
                compacted.append(stripped)
            continue
        compacted.append(value)
    return compacted


def parse_number(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if not text:
        return None

    lowered = text.lower()
    if lowered in {"n/a", "na", "not applicable", "xxxxx"}:
        return None

    cleaned = (
        text.replace("$", "")
        .replace(",", "")
        .replace("%", "")
        .replace("\u2212", "-")
        .replace("\xa0", "")
        .strip()
    )
    if not cleaned:
        return None

    try:
        if re.fullmatch(r"-?\d+", cleaned):
            return int(cleaned)
        return float(cleaned)
    except ValueError:
        return None


def parse_int(value: Any) -> int:
    number = parse_number(value)
    if number is None:
        raise ValueError(f"Could not parse integer from {value!r}")
    return int(round(float(number)))


def parse_rate(value: Any) -> float:
    number = parse_number(value)
    if number is None:
        return 0.0
    rate = float(number)
    return rate / 100 if rate > 1 else rate


def row_numbers(row: list[Any]) -> list[float | int]:
    values: list[float | int] = []
    for cell in row:
        number = parse_number(cell)
        if number is not None:
            values.append(number)
    return values


def midpoint(low: int, high: int) -> int:
    return int(round((low + high) / 2))


def weighted_average(values: list[float], weights: list[int]) -> int:
    total_weight = sum(weights)
    if total_weight == 0:
        return 0
    return int(round(sum(value * weight for value, weight in zip(values, weights)) / total_weight))


def read_xlsx_rows(path: Path, sheet_name: str) -> list[list[Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_map = {normalize_text(name): name for name in workbook.sheetnames}
        actual_name = sheet_map[normalize_text(sheet_name)]
        sheet = workbook[actual_name]
        return [compact_row(list(row)) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def find_row(rows: list[list[Any]], *needles: str) -> list[Any]:
    normalized_needles = [normalize_text(needle) for needle in needles]
    for row in rows:
        haystack = normalize_text(" | ".join(str(cell) for cell in row))
        if all(needle in haystack for needle in normalized_needles):
            return row
    raise ValueError(f"Could not find row containing {needles}")


def find_optional_row(rows: list[list[Any]], *needles: str) -> list[Any] | None:
    try:
        return find_row(rows, *needles)
    except ValueError:
        return None


def find_numeric_row(rows: list[list[Any]], *needles: str) -> list[Any]:
    normalized_needles = [normalize_text(needle) for needle in needles]
    for row in rows:
        haystack = normalize_text(" | ".join(str(cell) for cell in row))
        if all(needle in haystack for needle in normalized_needles) and row_numbers(row):
            return row
    raise ValueError(f"Could not find numeric row containing {needles}")


def find_optional_numeric_row(rows: list[list[Any]], *needles: str) -> list[Any] | None:
    try:
        return find_numeric_row(rows, *needles)
    except ValueError:
        return None


def extract_triplet(row: list[Any] | None) -> tuple[int, int, int]:
    numbers = [int(round(float(value))) for value in row_numbers(row or [])]
    if len(numbers) >= 3:
        return numbers[-3], numbers[-2], numbers[-1]
    if len(numbers) >= 2:
        low, high = numbers[-2], numbers[-1]
        return low, midpoint(low, high), high
    raise ValueError(f"Not enough numeric values in row: {row}")


def extract_total_from_race_row(row: list[Any]) -> int:
    numbers = row_numbers(row)
    if len(numbers) >= 2:
        return int(round(float(numbers[1])))
    if numbers:
        return int(round(float(numbers[-1])))
    raise ValueError(f"No numeric values found in row: {row}")


def build_by_race(rows: list[list[Any]]) -> dict[str, int]:
    label_map = {
        "international": ("nonresident aliens", "nonresidents", "international (nonresidents)"),
        "hispanicLatino": ("hispanic/latino",),
        "blackAfricanAmerican": ("black or african american",),
        "white": ("white, non-hispanic",),
        "asian": ("asian, non-hispanic",),
        "americanIndianAlaskaNative": ("american indian or alaska native",),
        "nativeHawaiianPacificIslander": ("native hawaiian or other pacific islander",),
        "twoOrMoreRaces": ("two or more races",),
        "unknown": ("race and/or ethnicity unknown",),
    }

    by_race: dict[str, int] = {}
    for key, labels in label_map.items():
        row = None
        for label in labels:
            row = find_optional_numeric_row(rows, label)
            if row is not None:
                break
        by_race[key] = extract_total_from_race_row(row) if row is not None else 0
    return by_race


def extract_xlsx_year(path: Path) -> dict[str, Any]:
    b_rows = read_xlsx_rows(path, "CDS-B")
    c_rows = read_xlsx_rows(path, "CDS-C")
    g_rows = read_xlsx_rows(path, "CDS-G")
    h_rows = read_xlsx_rows(path, "CDS-H")

    total_applied_row = find_optional_numeric_row(
        c_rows, "total first-time, first-year", "degree seeking", "applied"
    )
    total_admitted_row = find_optional_numeric_row(
        c_rows, "total first-time, first-year", "degree seeking", "admitted"
    )
    total_enrolled_row = find_optional_numeric_row(
        c_rows, "total first-time, first-year", "degree seeking", "enrolled"
    )

    if total_applied_row and total_admitted_row and total_enrolled_row:
        applied = parse_int(row_numbers(total_applied_row)[-1])
        admitted = parse_int(row_numbers(total_admitted_row)[-1])
        enrolled = parse_int(row_numbers(total_enrolled_row)[-1])
    else:
        applied = parse_int(row_numbers(find_row(c_rows, "men who applied"))[-1]) + parse_int(
            row_numbers(find_row(c_rows, "women who applied"))[-1]
        )
        admitted = parse_int(row_numbers(find_row(c_rows, "men who were admitted"))[-1]) + parse_int(
            row_numbers(find_row(c_rows, "women who were admitted"))[-1]
        )
        enrolled = 0
        for label in [
            "full-time, first-time, first-year (freshman) men who enrolled",
            "part-time, first-time, first-year (freshman) men who enrolled",
            "full-time, first-time, first-year (freshman) women who enrolled",
            "part-time, first-time, first-year (freshman) women who enrolled",
            "full-time, first-time, first-year men who enrolled",
            "part-time, first-time, first-year men who enrolled",
            "full-time, first-time, first-year women who enrolled",
            "part-time, first-time, first-year women who enrolled",
        ]:
            row = find_optional_row(c_rows, label)
            if row is not None:
                enrolled += parse_int(row_numbers(row)[-1])

    sat_submission_row = find_optional_row(c_rows, "submitting sat scores") or find_optional_row(
        c_rows, "percent submitting sat scores"
    )
    act_submission_row = find_optional_row(c_rows, "submitting act scores") or find_optional_row(
        c_rows, "percent submitting act scores"
    )
    sat_composite_row = find_optional_row(c_rows, "sat composite")
    sat_rw_row = find_row(c_rows, "sat evidence-based reading and writing")
    sat_math_row = find_row(c_rows, "sat math")
    act_composite_row = find_row(c_rows, "act composite")

    sat_rw = extract_triplet(sat_rw_row)
    sat_math = extract_triplet(sat_math_row)
    act_comp = extract_triplet(act_composite_row)

    if sat_composite_row is not None:
        sat_comp = extract_triplet(sat_composite_row)
    else:
        sat_comp = (
            sat_rw[0] + sat_math[0],
            midpoint(sat_rw[0] + sat_math[0], sat_rw[2] + sat_math[2]),
            sat_rw[2] + sat_math[2],
        )

    undergrad_row = find_optional_row(b_rows, "total undergraduate students") or find_row(
        b_rows, "total undergraduates"
    )
    graduate_row = find_optional_row(b_rows, "total graduate students") or find_row(b_rows, "total graduate")
    undergraduate = sum(parse_int(value) for value in row_numbers(undergrad_row))
    graduate = sum(parse_int(value) for value in row_numbers(graduate_row))

    by_race = build_by_race(b_rows)

    tuition_row = find_optional_numeric_row(g_rows, "tuition", "in-state") or find_optional_numeric_row(
        g_rows, "in-state (out-of-district)"
    )
    if tuition_row is None:
        raise ValueError(f"Missing in-state tuition row in {path.name}")
    fees_row = find_numeric_row(g_rows, "required fees")
    room_board_row = find_optional_numeric_row(g_rows, "room and board (on-campus)") or find_optional_numeric_row(
        g_rows, "room and board:", "(on-campus)"
    )
    if room_board_row is None:
        room_board_row = find_numeric_row(g_rows, "food and housing (on-campus)")

    aid_a_row = find_row(h_rows, "number of degree-seeking undergraduate students")
    aid_d_row = find_row(h_rows, "awarded any financial aid")
    aid_e_row = find_row(h_rows, "awarded any need-based scholarship")
    aid_h_row = find_row(h_rows, "need was fully met")
    aid_j_row = find_row(h_rows, "average financial aid package")
    aid_k_row = find_row(h_rows, "average need-based scholarship")

    a_numbers = [parse_int(value) for value in row_numbers(aid_a_row)]
    d_numbers = [parse_int(value) for value in row_numbers(aid_d_row)]
    e_numbers = [parse_int(value) for value in row_numbers(aid_e_row)]
    h_numbers = [parse_int(value) for value in row_numbers(aid_h_row)]
    j_numbers = [float(value) for value in row_numbers(aid_j_row)]
    k_numbers = [float(value) for value in row_numbers(aid_k_row)]

    total_students_for_aid = a_numbers[-2] + a_numbers[-1]
    total_students_with_aid = d_numbers[-2] + d_numbers[-1]
    total_students_with_need_grants = e_numbers[-2] + e_numbers[-1]
    total_students_need_met = h_numbers[-2] + h_numbers[-1]

    average_aid_package = weighted_average(j_numbers[-2:], d_numbers[-2:])
    average_need_based_grant = weighted_average(k_numbers[-2:], e_numbers[-2:])

    international = by_race["international"]

    return {
        "admissions": {
            "applied": applied,
            "admitted": admitted,
            "enrolled": enrolled,
            "acceptanceRate": round(admitted / applied, 4),
            "yield": round(enrolled / admitted, 4),
        },
        "testScores": {
            "sat": {
                "composite": {"p25": sat_comp[0], "p50": sat_comp[1], "p75": sat_comp[2]},
                "readingWriting": {"p25": sat_rw[0], "p50": sat_rw[1], "p75": sat_rw[2]},
                "math": {"p25": sat_math[0], "p50": sat_math[1], "p75": sat_math[2]},
                "submissionRate": parse_rate(row_numbers(sat_submission_row or [0])[0]),
            },
            "act": {
                "composite": {"p25": act_comp[0], "p50": act_comp[1], "p75": act_comp[2]},
                "submissionRate": parse_rate(row_numbers(act_submission_row or [0])[0]),
            },
        },
        "demographics": {
            "enrollment": {
                "total": undergraduate + graduate,
                "undergraduate": undergraduate,
                "graduate": graduate,
            },
            "byRace": by_race,
            "byResidency": {
                "inState": 0,
                "outOfState": 0,
                "international": international,
            },
        },
        "costs": {
            "tuition": parse_int(row_numbers(tuition_row)[0]),
            "fees": parse_int(row_numbers(fees_row)[0]),
            "roomAndBoard": parse_int(row_numbers(room_board_row)[0]),
            "totalCOA": parse_int(row_numbers(tuition_row)[0])
            + parse_int(row_numbers(fees_row)[0])
            + parse_int(row_numbers(room_board_row)[0]),
        },
        "financialAid": {
            "percentReceivingAid": round(total_students_with_aid / total_students_for_aid, 4),
            "averageAidPackage": average_aid_package,
            "averageNeedBasedGrant": average_need_based_grant,
            "percentNeedFullyMet": round(total_students_need_met / total_students_with_aid, 4),
        },
    }


def field_number(fields: dict[str, Any], key: str) -> float | int | None:
    field = fields.get(key)
    if field is None:
        return None
    return parse_number(field.get("/V"))


def field_int(fields: dict[str, Any], key: str) -> int:
    value = field_number(fields, key)
    if value is None:
        raise ValueError(f"Missing numeric PDF field {key}")
    return int(round(float(value)))


def field_rate(fields: dict[str, Any], key: str) -> float:
    return parse_rate(field_number(fields, key))


def extract_pdf_year(path: Path) -> dict[str, Any]:
    fields = PdfReader(str(path)).get_fields() or {}

    applied = field_int(fields, "AP_RECD_1ST_N")
    admitted = field_int(fields, "AP_ADMT_1ST_N")
    enrolled = field_int(fields, "EN_TOT_1ST_N")

    undergraduate = field_int(fields, "EN_TOT_UG_N")
    graduate = field_int(fields, "EN_TOT_GRAD_N")
    international = field_int(fields, "EN_TOT_NONRES_ALIEN_TOT_N")

    aid_total_students = field_int(fields, "UG_FT_N_N") + field_int(fields, "UG_PT_N_N")
    aid_recipients = field_int(fields, "UG_FT_REC_AID_N") + field_int(fields, "UG_PT_REC_AID")
    aid_need_grant_students = field_int(fields, "UG_FT_NB_GIFT_N") + field_int(fields, "UG_PT_NB_GIFT_N")
    aid_need_met_students = field_int(fields, "UG_FT_ND_MET_N") + field_int(fields, "UG_PT_ND_MET_N")

    average_aid_package = weighted_average(
        [field_int(fields, "UG_FT_AVG_PKG_D"), field_int(fields, "UG_PT_AVG_PKG_D")],
        [field_int(fields, "UG_FT_REC_AID_N"), field_int(fields, "UG_PT_REC_AID")],
    )
    average_need_based_grant = weighted_average(
        [field_int(fields, "UG_FT_AVG_NB_GIFT_D"), field_int(fields, "UG_PT_AVG_NB_GIFT_D")],
        [field_int(fields, "UG_FT_NB_GIFT_N"), field_int(fields, "UG_PT_NB_GIFT_N")],
    )

    room_and_board = field_int(fields, "RM_BD_D")
    tuition = field_int(fields, "TUIT_STATE_1ST_FT_D")
    fees = field_int(fields, "FEES_1ST_D")

    return {
        "admissions": {
            "applied": applied,
            "admitted": admitted,
            "enrolled": enrolled,
            "acceptanceRate": round(admitted / applied, 4),
            "yield": round(enrolled / admitted, 4),
        },
        "testScores": {
            "sat": {
                "composite": {
                    "p25": field_int(fields, "SAT1_COMP_25TH_P"),
                    "p50": field_int(fields, "SAT1_COMP_50TH_P"),
                    "p75": field_int(fields, "SAT1_COMP_75TH_P"),
                },
                "readingWriting": {
                    "p25": field_int(fields, "SAT1_VERB_25TH_P"),
                    "p50": field_int(fields, "SAT1_VERB_50TH_P"),
                    "p75": field_int(fields, "SAT1_VERB_75TH_P"),
                },
                "math": {
                    "p25": field_int(fields, "SAT1_MATH_25TH_P"),
                    "p50": field_int(fields, "SAT1_MATH_50TH_P"),
                    "p75": field_int(fields, "SAT1_MATH_75TH_P"),
                },
                "submissionRate": field_rate(fields, "SUBMIT_SAT1_P"),
            },
            "act": {
                "composite": {
                    "p25": field_int(fields, "ACT_COMP_25TH_P"),
                    "p50": field_int(fields, "ACT_COMP_50TH_P"),
                    "p75": field_int(fields, "ACT_COMP_75TH_P"),
                },
                "submissionRate": field_rate(fields, "SUBMIT_ACT_P"),
            },
        },
        "demographics": {
            "enrollment": {
                "total": undergraduate + graduate,
                "undergraduate": undergraduate,
                "graduate": graduate,
            },
            "byRace": {
                "international": international,
                "hispanicLatino": field_int(fields, "EN_TOT_HISPANIC_ETHNICITY_N"),
                "blackAfricanAmerican": field_int(fields, "EN_TOT_BLACK_NONHISPANIC_N"),
                "white": field_int(fields, "EN_TOT_WHITE_NONHISPANIC_N"),
                "asian": field_int(fields, "EN_TOT_ASIAN_NONHISPANIC_N"),
                "americanIndianAlaskaNative": field_int(fields, "EN_TOT_NATIVE_NONHISPANIC_N"),
                "nativeHawaiianPacificIslander": field_int(fields, "EN_TOT_ISLANDER_NONHISPANIC_N"),
                "twoOrMoreRaces": field_int(fields, "EN_TOT_MULTIRACE_NONHISPANIC_N"),
                "unknown": field_int(fields, "EN_TOT_RACE_ETHNICITY_UNKNOWN_N"),
            },
            "byResidency": {
                "inState": 0,
                "outOfState": 0,
                "international": international,
            },
        },
        "costs": {
            "tuition": tuition,
            "fees": fees,
            "roomAndBoard": room_and_board,
            "totalCOA": tuition + fees + room_and_board,
        },
        "financialAid": {
            "percentReceivingAid": round(aid_recipients / aid_total_students, 4),
            "averageAidPackage": average_aid_package,
            "averageNeedBasedGrant": average_need_based_grant,
            "percentNeedFullyMet": round(aid_need_met_students / aid_recipients, 4),
        },
    }


def main() -> None:
    years: dict[str, Any] = {}

    for year, filename in XLSX_FILES.items():
        years[year] = extract_xlsx_year(SOURCE_DIR / filename)

    for year, filename in PDF_FILES.items():
        years[year] = extract_pdf_year(SOURCE_DIR / filename)

    data = {
        "name": "University of Washington",
        "slug": "uwashington",
        "years": years,
    }

    OUTPUT_PATH.write_text(json.dumps(data, indent=2) + "\n")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
