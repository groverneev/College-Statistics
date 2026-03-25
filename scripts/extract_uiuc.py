#!/usr/bin/env python3

from __future__ import annotations

import argparse
import json
import math
import re
from pathlib import Path
from typing import Any

import openpyxl
import xlrd


def normalize_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value)).strip().lower()


def compact_row(values: list[Any]) -> list[Any]:
    compacted: list[Any] = []
    for value in values:
        if value is None:
            continue
        if isinstance(value, str):
            stripped = value.strip()
            if not stripped:
                continue
            compacted.append(stripped)
        else:
            compacted.append(value)
    return compacted


def parse_number(value: Any) -> float | int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if not text:
        return None

    lowered = text.lower()
    if lowered in {"n/a", "na", "not applicable", "varies", "xxxxx"}:
        return None

    cleaned = (
        text.replace("$", "")
        .replace(",", "")
        .replace("%", "")
        .replace("\u2212", "-")
        .strip()
    )
    if not cleaned:
        return None

    try:
        if re.fullmatch(r"-?\d+", cleaned):
            return int(cleaned)
        return float(cleaned)
    except ValueError:
        return None


def parse_rate(value: Any) -> float:
    number = parse_number(value)
    if number is None:
        return 0.0
    rate = float(number)
    return rate / 100 if rate > 1 else rate


def row_numbers(row: list[Any]) -> list[float | int]:
    numbers: list[float | int] = []
    for cell in row:
        number = parse_number(cell)
        if number is not None:
            numbers.append(number)
    return numbers


def midpoint(p25: int, p75: int) -> int:
    return int(round((p25 + p75) / 2))


def read_xlsx_rows(path: Path, sheet_name: str) -> list[list[Any]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        sheet_map = {normalize_text(name): name for name in workbook.sheetnames}
        actual_name = sheet_map.get(normalize_text(sheet_name))
        if actual_name is None:
            raise ValueError(f"Sheet {sheet_name!r} not found in {path.name}")
        sheet = workbook[actual_name]
        return [compact_row(list(row)) for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def read_xls_rows(path: Path, sheet_name: str) -> list[list[Any]]:
    workbook = xlrd.open_workbook(path)
    sheet_map = {normalize_text(name): name for name in workbook.sheet_names()}
    actual_name = sheet_map.get(normalize_text(sheet_name))
    if actual_name is None:
        raise ValueError(f"Sheet {sheet_name!r} not found in {path.name}")
    sheet = workbook.sheet_by_name(actual_name)

    rows: list[list[Any]] = []
    for row_idx in range(sheet.nrows):
        rows.append(compact_row(sheet.row_values(row_idx)))
    return rows


def read_rows(path: Path, sheet_name: str) -> list[list[Any]]:
    if path.suffix.lower() == ".xlsx":
        return read_xlsx_rows(path, sheet_name)
    if path.suffix.lower() == ".xls":
        return read_xls_rows(path, sheet_name)
    raise ValueError(f"Unsupported workbook type: {path.suffix}")


def find_matching_rows(rows: list[list[Any]], *needles: str) -> list[list[Any]]:
    normalized_needles = [normalize_text(needle) for needle in needles]
    matches: list[list[Any]] = []
    for row in rows:
        haystack = normalize_text(" | ".join(str(cell) for cell in row))
        if all(needle in haystack for needle in normalized_needles):
            matches.append(row)
    return matches


def find_row(rows: list[list[Any]], *needles: str) -> list[Any]:
    matches = find_matching_rows(rows, *needles)
    if not matches:
        raise ValueError(f"Could not find row containing: {needles}")
    return matches[0]


def find_optional_row(rows: list[list[Any]], *needles: str) -> list[Any] | None:
    matches = find_matching_rows(rows, *needles)
    return matches[0] if matches else None


def find_numeric_row(rows: list[list[Any]], *needles: str) -> list[Any]:
    for row in find_matching_rows(rows, *needles):
        if row_numbers(row):
            return row
    raise ValueError(f"Could not find numeric row containing: {needles}")


def find_optional_numeric_row(rows: list[list[Any]], *needles: str) -> list[Any] | None:
    try:
        return find_numeric_row(rows, *needles)
    except ValueError:
        return None


def find_numeric_rows(rows: list[list[Any]], *needles: str) -> list[list[Any]]:
    return [row for row in find_matching_rows(rows, *needles) if row_numbers(row)]


def extract_last_int(row: list[Any]) -> int:
    numbers = row_numbers(row)
    if not numbers:
        raise ValueError(f"No numeric values found in row: {row}")
    return int(round(float(numbers[-1])))


def select_undergraduate_value(rows: list[list[Any]], *, rate: bool = False) -> int | float:
    if not rows:
        raise ValueError("No rows provided for undergraduate value selection")

    numeric_rows = [row for row in rows if row_numbers(row)]
    if not numeric_rows:
        raise ValueError("No numeric rows provided for undergraduate value selection")

    first_numbers = row_numbers(numeric_rows[0])
    if len(first_numbers) >= 2:
        value = first_numbers[1]
    elif len(numeric_rows) >= 2:
        value = row_numbers(numeric_rows[1])[0]
    else:
        value = first_numbers[0]

    return parse_rate(value) if rate else int(round(float(value)))


def select_total_undergrad_b2_value(rows: list[list[Any]]) -> int:
    if not rows:
        raise ValueError("No rows provided for B2 value selection")

    numeric_rows = [row for row in rows if row_numbers(row)]
    if not numeric_rows:
        raise ValueError("No numeric rows provided for B2 value selection")

    first_numbers = row_numbers(numeric_rows[0])
    if len(first_numbers) >= 2:
        value = first_numbers[-1]
    else:
        value = row_numbers(numeric_rows[-1])[0]

    return int(round(float(value)))


def extract_percentile_triplet(
    rows: list[list[Any]],
    percentile_needles: tuple[str, ...],
    combined_needles: tuple[str, ...],
) -> tuple[int, int, int]:
    percentile_rows = [
        find_optional_numeric_row(rows, *percentile_needles, "25th percentile"),
        find_optional_numeric_row(rows, *percentile_needles, "50th percentile"),
        find_optional_numeric_row(rows, *percentile_needles, "75th percentile"),
    ]
    if all(row is not None for row in percentile_rows):
        return tuple(extract_last_int(row) for row in percentile_rows)  # type: ignore[return-value]

    combined_row = find_numeric_row(rows, *combined_needles)
    numbers = [int(round(float(value))) for value in row_numbers(combined_row)]
    if len(numbers) >= 3:
        return numbers[-3], numbers[-2], numbers[-1]
    if len(numbers) >= 2:
        p25, p75 = numbers[-2], numbers[-1]
        return p25, midpoint(p25, p75), p75
    raise ValueError(f"Not enough score values in row: {combined_row}")


def extract_admissions(rows: list[list[Any]]) -> dict[str, Any]:
    aggregate_applied = find_optional_numeric_row(
        rows, "total first-time, first-year students", "applied"
    )
    aggregate_admitted = find_optional_numeric_row(
        rows, "total first-time, first-year students", "admitted"
    )
    aggregate_enrolled = find_optional_numeric_row(
        rows, "total first-time, first-year students", "enrolled"
    )

    if aggregate_applied and aggregate_admitted and aggregate_enrolled:
        applied = sum(int(round(float(value))) for value in row_numbers(aggregate_applied))
        admitted = sum(int(round(float(value))) for value in row_numbers(aggregate_admitted))
        enrolled = sum(int(round(float(value))) for value in row_numbers(aggregate_enrolled))
    else:
        applied = sum(
            extract_last_int(find_numeric_row(rows, "total", label, "applied"))
            for label in ("men", "women")
        )
        admitted = sum(
            extract_last_int(find_numeric_row(rows, "total", label, "admitted"))
            for label in ("men", "women")
        )
        enrolled = 0
        for labels in [
            ("total", "enrolled", "full-time", "men"),
            ("total", "enrolled", "part-time", "men"),
            ("total", "enrolled", "full-time", "women"),
            ("total", "enrolled", "part-time", "women"),
            ("total", "enrolled", "full-time", "unknown gender"),
            ("total", "enrolled", "part-time", "unknown gender"),
            ("total", "enrolled", "full-time", "another gender"),
            ("total", "enrolled", "part-time", "another gender"),
        ]:
            row = find_optional_numeric_row(rows, *labels)
            if row is not None and row_numbers(row):
                enrolled += extract_last_int(row)

    data: dict[str, Any] = {
        "applied": applied,
        "admitted": admitted,
        "enrolled": enrolled,
        "acceptanceRate": round(admitted / applied, 4),
        "yield": round(enrolled / admitted, 4),
    }

    early_applied = find_optional_numeric_row(rows, "number of early decision applications received")
    early_admitted = find_optional_numeric_row(rows, "admitted under early decision")
    if early_applied is not None and early_admitted is not None:
        data["earlyDecision"] = {
            "applied": extract_last_int(early_applied),
            "admitted": extract_last_int(early_admitted),
        }

    return data


def extract_test_scores(rows: list[list[Any]]) -> dict[str, Any]:
    sat_submission_row = (
        find_optional_row(rows, "submitting sat scores")
        or find_optional_row(rows, "percent submitting sat scores")
    )
    act_submission_row = (
        find_optional_row(rows, "submitting act scores")
        or find_optional_row(rows, "percent submitting act scores")
    )

    sat_rw_p25, sat_rw_p50, sat_rw_p75 = extract_percentile_triplet(
        rows,
        ("sat evidence-based reading and writing",),
        ("sat evidence-based reading and writing",),
    ) if find_optional_numeric_row(rows, "sat evidence-based reading and writing") or find_optional_numeric_row(
        rows, "sat evidence-based reading and writing", "25th percentile"
    ) else extract_percentile_triplet(
        rows,
        ("sat critical reading",),
        ("sat critical reading",),
    )
    sat_math_p25, sat_math_p50, sat_math_p75 = extract_percentile_triplet(
        rows,
        ("sat math",),
        ("sat math",),
    )
    sat_composite_row = find_optional_numeric_row(rows, "sat composite")
    if sat_composite_row is not None or find_optional_numeric_row(rows, "sat composite", "25th percentile") is not None:
        sat_comp_p25, sat_comp_p50, sat_comp_p75 = extract_percentile_triplet(
            rows,
            ("sat composite",),
            ("sat composite",),
        )
    else:
        sat_comp_p25 = sat_rw_p25 + sat_math_p25
        sat_comp_p50 = sat_rw_p50 + sat_math_p50
        sat_comp_p75 = sat_rw_p75 + sat_math_p75
    act_p25, act_p50, act_p75 = extract_percentile_triplet(
        rows,
        ("act composite",),
        ("act composite",),
    )

    sat_submission_rate = parse_rate(row_numbers(sat_submission_row)[0]) if sat_submission_row else 0.0
    act_submission_rate = parse_rate(row_numbers(act_submission_row)[0]) if act_submission_row else 0.0

    return {
        "sat": {
            "composite": {"p25": sat_comp_p25, "p50": sat_comp_p50, "p75": sat_comp_p75},
            "readingWriting": {"p25": sat_rw_p25, "p50": sat_rw_p50, "p75": sat_rw_p75},
            "math": {"p25": sat_math_p25, "p50": sat_math_p50, "p75": sat_math_p75},
            "submissionRate": round(sat_submission_rate, 4),
        },
        "act": {
            "composite": {"p25": act_p25, "p50": act_p50, "p75": act_p75},
            "submissionRate": round(act_submission_rate, 4),
        },
    }


def extract_demographics(b_rows: list[list[Any]], f_rows: list[list[Any]]) -> dict[str, Any]:
    undergraduate_row = (
        find_optional_numeric_row(b_rows, "total all undergraduates")
        or find_optional_numeric_row(b_rows, "total of all undergraduate students enrolled")
        or find_numeric_row(b_rows, "total undergraduate students")
    )
    graduate_row = (
        find_optional_numeric_row(b_rows, "total all graduate")
        or find_optional_numeric_row(b_rows, "total of all graduate students enrolled")
        or find_numeric_row(b_rows, "total graduate students")
    )
    total_row = find_optional_numeric_row(b_rows, "grand total all students")

    undergraduate = extract_last_int(undergraduate_row)
    graduate = extract_last_int(graduate_row)
    total = extract_last_int(total_row) if total_row is not None else undergraduate + graduate

    by_race = {
        "international": select_total_undergrad_b2_value(
            find_numeric_rows(b_rows, "nonresident aliens") or find_numeric_rows(b_rows, "nonresidents")
        ),
        "hispanicLatino": select_total_undergrad_b2_value(find_numeric_rows(b_rows, "hispanic/latino")),
        "blackAfricanAmerican": select_total_undergrad_b2_value(
            find_numeric_rows(b_rows, "black or african american, non-hispanic")
        ),
        "white": select_total_undergrad_b2_value(find_numeric_rows(b_rows, "white, non-hispanic")),
        "asian": select_total_undergrad_b2_value(find_numeric_rows(b_rows, "asian, non-hispanic")),
        "americanIndianAlaskaNative": select_total_undergrad_b2_value(
            find_numeric_rows(b_rows, "american indian or alaska native, non-hispanic")
        ),
        "nativeHawaiianPacificIslander": select_total_undergrad_b2_value(
            find_numeric_rows(b_rows, "native hawaiian or other pacific islander, non-hispanic")
        ),
        "twoOrMoreRaces": select_total_undergrad_b2_value(
            find_numeric_rows(b_rows, "two or more races, non-hispanic")
        ),
        "unknown": select_total_undergrad_b2_value(find_numeric_rows(b_rows, "race and/or ethnicity unknown")),
    }

    out_of_state_rows = find_numeric_rows(f_rows, "out of state")
    out_of_state_pct = select_undergraduate_value(out_of_state_rows, rate=True)
    domestic = undergraduate - by_race["international"]
    out_of_state = int(math.floor(domestic * out_of_state_pct + 0.5))
    in_state = domestic - out_of_state

    return {
        "enrollment": {
            "total": total,
            "undergraduate": undergraduate,
            "graduate": graduate,
        },
        "byRace": by_race,
        "byResidency": {
            "inState": in_state,
            "outOfState": out_of_state,
            "international": by_race["international"],
        },
    }


def extract_costs(rows: list[list[Any]]) -> dict[str, int]:
    tuition_row = (
        find_optional_numeric_row(rows, "tuition: in-state (out-of-district)")
        or find_optional_numeric_row(rows, "public institutions", "tuition", "in-district")
        or find_optional_numeric_row(rows, "tuition: in-district")
        or find_numeric_row(rows, "tuition:")
    )
    fees_row = find_numeric_row(rows, "required fees")
    room_board_row = (
        find_optional_numeric_row(rows, "food and housing", "on-campus")
        or find_optional_numeric_row(rows, "room and board", "on-campus")
        or find_optional_numeric_row(rows, "room and board")
    )
    if room_board_row is None:
        raise ValueError("Missing room and board row")

    tuition = int(round(float(row_numbers(tuition_row)[0])))
    fees = int(round(float(row_numbers(fees_row)[0])))
    room_and_board = int(round(float(row_numbers(room_board_row)[0])))

    return {
        "tuition": tuition,
        "fees": fees,
        "roomAndBoard": room_and_board,
        "totalCOA": tuition + fees + room_and_board,
    }


def extract_financial_aid(rows: list[list[Any]]) -> dict[str, Any]:
    total_students = select_undergraduate_value(
        find_numeric_rows(rows, "number of degree-seeking undergraduate students")
    )
    awarded_aid = select_undergraduate_value(find_numeric_rows(rows, "awarded any financial aid"))
    fully_met_rate = select_undergraduate_value(
        find_numeric_rows(rows, "percentage of need that was met"),
        rate=True,
    )
    avg_package = select_undergraduate_value(find_numeric_rows(rows, "average financial aid package"))
    avg_grant = select_undergraduate_value(
        find_numeric_rows(rows, "average need-based scholarship", "grant")
    )

    return {
        "percentReceivingAid": round(awarded_aid / total_students, 4),
        "averageAidPackage": avg_package,
        "averageNeedBasedGrant": avg_grant,
        "percentNeedFullyMet": round(fully_met_rate, 4),
    }


def extract_year_data(workbook_path: Path) -> dict[str, Any]:
    admissions_rows = read_rows(workbook_path, "CDS-C")
    demographics_rows = read_rows(workbook_path, "CDS-B")
    student_life_rows = read_rows(workbook_path, "CDS-F")
    costs_rows = read_rows(workbook_path, "CDS-G")
    aid_rows = read_rows(workbook_path, "CDS-H")

    year_data = {
        "admissions": extract_admissions(admissions_rows),
        "testScores": extract_test_scores(admissions_rows),
        "demographics": extract_demographics(demographics_rows, student_life_rows),
        "costs": extract_costs(costs_rows),
        "financialAid": extract_financial_aid(aid_rows),
    }

    undergrad = year_data["demographics"]["enrollment"]["undergraduate"]
    race_total = sum(year_data["demographics"]["byRace"].values())
    residency_total = sum(year_data["demographics"]["byResidency"].values())
    if race_total != undergrad:
        raise ValueError(
            f"{workbook_path.name}: race total {race_total} does not match undergrad {undergrad}"
        )
    if residency_total != undergrad:
        raise ValueError(
            f"{workbook_path.name}: residency total {residency_total} does not match undergrad {undergrad}"
        )

    return year_data


def workbook_year_label(path: Path) -> str:
    match = re.search(r"(20\d{2})[-_](20\d{2})", path.stem)
    if not match:
        raise ValueError(f"Could not determine year from filename: {path.name}")
    return f"{match.group(1)}-{match.group(2)}"


def build_dataset(input_dir: Path) -> dict[str, Any]:
    years: dict[str, Any] = {}
    for workbook_path in sorted(input_dir.iterdir()):
        if workbook_path.suffix.lower() not in {".xls", ".xlsx"}:
            continue
        years[workbook_year_label(workbook_path)] = extract_year_data(workbook_path)

    return {
        "name": "University of Illinois Urbana-Champaign",
        "slug": "uiuc",
        "years": years,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Extract UIUC CDS data from Excel workbooks.")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("College-Data/UIllinoisUrbanaChampaign"),
    )
    parser.add_argument("--output", type=Path, default=Path("src/data/schools/uiuc.json"))
    args = parser.parse_args()

    dataset = build_dataset(args.input_dir)
    args.output.write_text(json.dumps(dataset, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
